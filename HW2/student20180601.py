#!/usr/bin/python3
import openpyxl
import math

wb = openpyxl.load_workbook("HW2/student.xlsx")
ws = wb['Sheet1']

total = []
row_id = 1
for row in ws:
	if row_id != 1:
		sum_v = ws.cell(row = row_id, column = 3).value * 0.3
		sum_v += ws.cell(row = row_id, column = 4).value * 0.35
		sum_v += ws.cell(row = row_id, column = 5).value * 0.34
		sum_v += ws.cell(row = row_id, column = 6).value
		ws.cell(row = row_id, column = 7).value = sum_v
		total.append([row_id, sum_v])
	row_id += 1

total.sort(key=lambda x:x[1], reverse=True)
count = len(total)

#A
A_zero = int(count * 0.3)
A_plus = int(A_zero * 0.5)
A0 = A_zero - A_plus;

#B
B_zero = int(count * 0.7)
B_plus = int((B_zero - A_zero) * 0.5)
B0 = B_zero - A_zero - B_plus;

#C
C_zero = count
C_plus = int((C_zero - B_zero) * 0.5)
C0 = C_zero - B_zero - C_plus

for i in range (count):
        ws.cell(row = total[i][0], column = 8).value = 'C0'
for i in range(int(C_zero - C0)):
        ws.cell(row = total[i][0], column = 8).value = 'C+'
for i in range(int(B_zero)):
        ws.cell(row = total[i][0], column = 8).value = 'B0'
for i in range(int(B_zero - B0)):
        ws.cell(row = total[i][0], column = 8).value = 'B+'
for i in range(int(A_zero)):
        ws.cell(row = total[i][0], column = 8).value = 'A0'
for i in range(int(A_plus)):
        ws.cell(row = total[i][0], column = 8).value = 'A+'

wb.save("student.xlsx")